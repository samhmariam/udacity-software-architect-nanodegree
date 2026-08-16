"""Complete the supplied verification and validation test-case workbook."""

from pathlib import Path
from shutil import copy2

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


PROJECT = Path(__file__).resolve().parents[1]
SOURCE = PROJECT / "test-cases-export_sample.xlsx"
OUTPUT = PROJECT / "test-cases-export_completed.xlsx"

VERIFICATION = [
    {
        "id": "TC-V-001",
        "name": "Verify Motion Sensor Alerts",
        "priority": "Critical",
        "type": "Functional — Integration and timing",
        "objective": "Verify that an enrolled motion sensor reports authenticated movement and that the CCU evaluates it according to mode and zone policy.",
        "pre": "Production-representative CCU and motion sensor; sensor enrolled, named, assigned to an enabled instant interior zone, and showing healthy; system time synchronized; event-log access available; instrumented timestamps with ≤10 ms uncertainty.",
        "steps": [
            ("Place the system in Disarmed and inspect device status.", "Sensor is Online/Healthy and its configured name and zone are correct."),
            ("Create motion inside the specified detection area.", "Sensor detects motion and transmits an authenticated event."),
            ("Inspect the CCU event log and timestamps.", "Correct device ID, zone, event type, sequence/counter, and UTC timestamp are recorded; 99% of 100 trials register within 500 ms of transmission."),
            ("Arm Away, clear the exit delay, and create motion again.", "The enabled interior event becomes an intrusion candidate and the CCU enters Alarm."),
            ("Arm Stay with the same zone configured as Stay-bypassed and repeat.", "The event is logged but does not trigger Alarm."),
        ],
        "expected": "Pass if enrollment/status are correct, event data is authenticated and accurate, PR-01 timing passes, Armed-Away alarms, and configured Armed-Stay bypass prevents an alarm. Fail on a missed/incorrect event, timing breach, or wrong state transition.",
        "req": "FR-01, FR-04, FR-05, PR-01, IF-01",
    },
    {
        "id": "TC-V-002",
        "name": "Verify Door/Window Sensor Alerts",
        "priority": "Critical",
        "type": "Functional — Integration and timing",
        "objective": "Verify accurate open/closed reporting and instant versus delayed-entry behavior for an enrolled perimeter sensor.",
        "pre": "Door/window contact enrolled and healthy; magnet aligned; zone configurable as instant or delayed entry; entry delay set to 30 s; event-log and timing access available.",
        "steps": [
            ("In Disarmed, confirm the opening is physically closed and inspect status.", "UI reports Closed for the correct named opening."),
            ("Open and then close the opening while recording timestamps.", "Open and Closed events are registered with correct order and identity; 99% of trials register within 500 ms of sensor transmission."),
            ("Configure the zone as instant, arm Away, clear exit delay, and open it.", "CCU enters Alarm and commands the siren without entry delay."),
            ("Configure the zone as delayed entry, re-arm, clear exit delay, and open it.", "CCU enters Entry Delay and displays/counts the configured 30 s."),
            ("Disarm with an authorized PIN before expiry; repeat and allow the timer to expire.", "First attempt returns to Disarmed without Alarm; second enters Alarm at expiry."),
        ],
        "expected": "Pass if open/closed state, identity, order, timing, instant-zone alarm, and delayed-entry transitions all match the configured policy. Fail for a missed/reversed event or incorrect transition.",
        "req": "FR-01, FR-06, FR-07, PR-01, IF-01",
    },
    {
        "id": "TC-V-003",
        "name": "Verify Water Leak Sensor Integration",
        "priority": "High",
        "type": "Functional — Integration and recovery",
        "objective": "Verify that a supervised water-leak sensor reports wet/dry state and produces a distinct property-protection alert without being misclassified as intrusion.",
        "pre": "Water sensor enrolled, named, healthy, and assigned to a utility zone; absorbent test surface and manufacturer-approved test water available; mobile recipient configured; cleanup materials available.",
        "steps": [
            ("Inspect the device and zone status while the sensor is dry.", "Device is Healthy and state is Dry."),
            ("Apply the manufacturer-specified amount of water to the sensing area/cable.", "Sensor changes to Wet and sends an authenticated leak event."),
            ("Inspect the CCU log, local UI, and recipient alert.", "Each identifies Water Leak, residence/zone, sensor name, and timestamp; it is not labelled as intrusion."),
            ("Remove water, dry the sensor, and reset/acknowledge as specified.", "State returns to Dry/Ready; the historical event remains in the audit log."),
            ("Simulate low battery and then loss of heartbeat.", "Distinct health faults are detected and reported without deleting the leak history."),
        ],
        "expected": "Pass if wet/dry state, distinct leak notification, audit retention, and supervision operate correctly. Fail if the event is missed, misclassified, or health loss is silent.",
        "req": "FR-01, FR-10, FR-12, FR-15, IF-01",
    },
    {
        "id": "TC-V-004",
        "name": "View Camera Feed and Verify Camera Integration",
        "priority": "Critical",
        "type": "Functional — Integration; Non-functional — Performance",
        "objective": "Verify authenticated camera health, live access, and event-correlated video preservation around an alarm trigger.",
        "pre": "Camera enrolled, online, time-synchronized, and associated with the test zone; event recording enabled; sufficient encrypted storage; scene clock visible; authorized owner account available.",
        "steps": [
            ("Open the camera status and authorized local live view.", "Camera is Healthy and the correct live stream appears without exposing another camera."),
            ("Generate identifiable activity for at least 15 s before triggering the associated intrusion sensor.", "Camera pre-event buffer contains the activity."),
            ("Trigger the associated sensor and continue activity for at least 35 s.", "CCU creates an incident and requests/preserves the event clip."),
            ("Open the completed clip and compare it with the scene clock.", "Clip contains at least 10 s before and 30 s after the trigger with correct camera, zone, and incident linkage."),
            ("Remove camera network connectivity and trigger the sensor again.", "Camera fault is reported; local detection and siren still operate."),
        ],
        "expected": "Pass if live view is authorized, clip coverage meets FR-11, metadata is correct, and camera loss degrades evidence only—not local alarming. Fail on unauthorized access, insufficient clip duration, or suppressed alarm.",
        "req": "FR-08, FR-10, FR-11, RE-02, SEC-01, SEC-03",
    },
    {
        "id": "TC-V-005",
        "name": "Verify Login — Successful Authentication",
        "priority": "Critical",
        "type": "Functional — Authentication; Non-functional — Security",
        "objective": "Verify that valid local and remote credentials authenticate the correct identity and enforce its assigned least-privilege role.",
        "pre": "Active administrator, member, guest, and installer accounts; MFA configured for remote administrator; clean browser/app sessions; audit access through a separate administrator.",
        "steps": [
            ("Sign in locally with a valid member PIN.", "Member is authenticated and receives member functions only."),
            ("Sign out, then sign in remotely with valid administrator credentials and first factor.", "System requests the configured second factor and does not create an authenticated session yet."),
            ("Provide the valid second factor.", "Administrator session is created over TLS and privileged configuration is available."),
            ("Sign in with valid guest and installer credentials in separate sessions.", "Guest cannot configure/view recordings; installer access is time-limited and cannot view recordings without explicit temporary owner grant."),
            ("Review the security audit log.", "Each login contains identity, result, channel, and timestamp without recording secrets."),
        ],
        "expected": "Pass if valid users authenticate, remote admin requires MFA, every role is least-privileged, installer restrictions apply, and logins are audited. Fail for privilege leakage, MFA bypass, or exposed credentials.",
        "req": "FR-03, HS-03, SEC-01, SEC-02, SEC-06",
    },
    {
        "id": "TC-V-006",
        "name": "Verify Login — Failed Authentication",
        "priority": "Critical",
        "type": "Non-functional — Security and abuse resistance",
        "objective": "Verify that incorrect, replayed, expired, and malformed authentication attempts fail securely without revealing account information or disabling emergency functions.",
        "pre": "Known test account; audit access; ability to submit local PIN and remote token requests; emergency alarm function available; clocks synchronized.",
        "steps": [
            ("Submit an unknown username, a known username with wrong password, and an invalid PIN.", "All fail with equivalent non-enumerating messages and no authenticated session/state change."),
            ("Submit five wrong local PIN attempts within 10 min.", "Rate-limited lockout is imposed and recorded."),
            ("During lockout, attempt protected disarm and the emergency alarm function.", "Disarm is denied; emergency function remains available."),
            ("Replay a previously valid token/command and submit expired and malformed tokens.", "Each is rejected without state change or sensitive diagnostic disclosure."),
            ("Inspect audit events and wait for/configure lockout recovery.", "Failures, rate limit, and attack indicators are timestamped; authorized recovery restores access without clearing evidence."),
        ],
        "expected": "Pass if every invalid attempt fails closed, five-attempt lockout/rate limiting operates, emergency functions remain usable, and security events are audited. Fail on login, disarm, enumeration, secret leakage, or silent attack handling.",
        "req": "FR-03, SEC-02, SEC-06, SEC-08",
    },
    {
        "id": "TC-V-007",
        "name": "Verify Sensor Tamper Detection",
        "priority": "High",
        "type": "Non-functional — Security and reliability",
        "objective": "Verify that physical removal/opening and loss of sensor communications are detected, identified, reported, and retained as distinct health/security events.",
        "pre": "Tamper-capable sensor enrolled and healthy; case and removal tamper mechanisms enabled; local display and remote recipient configured; heartbeat timeout configured to the production value.",
        "steps": [
            ("Open the sensor enclosure without disarming its tamper mechanism.", "CCU reports Tamper for the correct sensor and records its timestamp."),
            ("Restore the enclosure, then remove the sensor from its mounting surface.", "Removal tamper is detected and history remains visible after restore."),
            ("Block/remove sensor communications while leaving the CCU online.", "CCU detects communication loss within 5 min and reports it locally within 1 additional min."),
            ("Restore communications and inspect health/audit views.", "Current health returns to Healthy only after a valid heartbeat; all tamper/loss/restore events remain ordered in the audit log."),
        ],
        "expected": "Pass if enclosure/removal tamper and heartbeat loss identify the correct device, satisfy RE-03 timing, recover safely, and remain audited. Fail if loss of protection is silent or history disappears.",
        "req": "FR-10, FR-15, RE-03, PH-02, SEC-06",
    },
]

VALIDATION = [
    {
        "id": "TC-UA-001",
        "suite": "Validation - User Acceptance",
        "name": "Motion Detection and Alert Notification",
        "priority": "Critical",
        "type": "Functional — User acceptance",
        "objective": "Validate that an owner away from home can understand and appropriately respond to a real-world motion alarm.",
        "pre": "Representative furnished residence; Armed-Away with exit delay complete; owner outside the home; notifications enabled; associated camera healthy; stable Internet/provider; no facilitator guidance after scenario begins.",
        "steps": [
            ("A test actor enters and moves through the protected motion zone.", "Local alarm activates and an incident is created."),
            ("Owner observes the received mobile alert.", "Provider submission occurs within 5 s for this nominal scenario; alert shows residence, zone, time, and actions on one screen."),
            ("Owner opens the incident and reviews available video.", "Correct event clip and current device status are understandable and associated with the same incident."),
            ("Owner acknowledges and silences the alarm after authentication.", "Siren stops, incident becomes Acknowledged, and evidence/audit history remains."),
        ],
        "expected": "Pass if the end-to-end scenario succeeds and the participant identifies the residence/zone and chooses an appropriate action without help; fail for missing/misleading alert, wrong evidence, or inability to respond.",
        "req": "FR-08, FR-09, FR-11, PR-03, UR-03, VAL-02",
    },
    {
        "id": "TC-UA-002",
        "suite": "Validation - User Acceptance",
        "name": "Door/Window Sensor Activation",
        "priority": "Critical",
        "type": "Functional — User acceptance",
        "objective": "Validate that an occupant returning through a delayed-entry door understands the countdown and can disarm safely, while an unattended entry produces an alarm.",
        "pre": "Representative entry door configured with 30 s delay; system Armed-Away; authorized participant returning home; local panel visible/audible; separate unattended-entry run planned.",
        "steps": [
            ("Participant opens the entry door.", "Entry Delay starts and panel clearly identifies the entry zone and remaining time."),
            ("Participant uses the local panel to disarm with a valid credential.", "System reaches Disarmed before expiry, with no siren alarm."),
            ("Re-arm and repeat the entry without disarming.", "At delay expiry, siren activates, incident is created, and recipients are alerted."),
            ("Ask the participant to explain both outcomes.", "Participant correctly distinguishes safe entry from intrusion alarm."),
        ],
        "expected": "Pass if the participant disarms without assistance before expiry, correctly explains state/zone, and the unattended run alarms at expiry. Fail for a critical use error or incorrect system response.",
        "req": "FR-07, FR-08, HS-01, HS-04, UR-02, VAL-01",
    },
    {
        "id": "TC-UA-003",
        "suite": "Validation - User Acceptance",
        "name": "Remote Camera Access and Live View",
        "priority": "High",
        "type": "Functional — User acceptance; Non-functional — Privacy",
        "objective": "Validate that an authorized remote owner can locate and use the intended camera while unauthorized roles and unavailable cameras are handled safely.",
        "pre": "Owner on an external network; member and guest test identities; at least two named cameras; one camera can be disconnected; valid remote MFA configured; consent/retention settings visible.",
        "steps": [
            ("Owner signs in remotely using MFA and selects a named camera.", "Correct live view opens and current synchronization/health status is visible."),
            ("Owner opens a recent event clip, then exports it.", "Correct clip is displayed/exported through an authenticated, encrypted interaction and the access is audited."),
            ("Guest attempts to open the same camera and clip.", "Access is denied without revealing media or sensitive metadata."),
            ("Disconnect the selected camera and refresh the owner view.", "UI shows a clear camera-specific unavailable/degraded state, not stale video presented as live."),
        ],
        "expected": "Pass if authorized live/recorded access is understandable and audited, guest access is denied, and outage status is explicit. Fail on wrong/stale media, privacy leakage, or misleading health.",
        "req": "FR-12, SEC-01, SEC-03, SEC-06, IM-05, VAL-02",
    },
    {
        "id": "TC-UA-004",
        "suite": "Validation - User Acceptance",
        "name": "System Arm and Disarm Functionality",
        "priority": "Critical",
        "type": "Functional — User acceptance; Non-functional — Usability",
        "objective": "Validate that first-time occupants can arm Away and disarm locally with correct feedback and without assistance.",
        "pre": "At least 20 first-time adult participants representative of intended users, including assistive-technology users; five-minute orientation; configured local panel and mobile app; timing observer; 100–500 lux test environment.",
        "steps": [
            ("Ask each participant to identify the current mode and any displayed blocking fault from 1 m.", "Participant correctly identifies mode/fault without relying on color alone."),
            ("Ask the participant to arm Away and leave through the designated route.", "Exit Delay is apparent; system reaches Armed-Away after the configured delay."),
            ("Ask the participant to return and disarm locally.", "Entry Delay is apparent; authorized disarm reaches Disarmed without an alarm."),
            ("Record completion time, assistance, errors, and satisfaction.", "Objective usability data is captured without coaching."),
        ],
        "expected": "Pass if ≥90% complete arm and disarm without assistance within 60 s after orientation, ≥95% identify mode/fault, accessibility checks pass, and no critical use error occurs.",
        "req": "FR-02, FR-03, UR-01, UR-02, UR-05, VAL-01",
    },
    {
        "id": "TC-LT-001",
        "suite": "Validation - Load Testing",
        "name": "Simultaneous Sensor Triggers",
        "priority": "Critical",
        "type": "Non-functional — Performance, load, and concurrency",
        "objective": "Validate that concurrent events from three or more sensors form one understandable incident without loss, duplication, or unsafe state behavior.",
        "pre": "Armed-Away; healthy motion sensor, entry sensor, and tamper-capable sensor in distinct named zones; synchronized event injector/actors; camera associations and recipients configured; instrumentation enabled.",
        "steps": [
            ("Trigger the three sensors within the same 100 ms window.", "CCU accepts all valid authenticated events and enters Alarm once."),
            ("Inspect the active incident, siren, camera requests, notifications, and audit log.", "All three events appear once with correct identity/order; siren is commanded; associated evidence and one coherent alert/incident are produced."),
            ("Repeat 100 times while the system carries nominal PR-05 device/session load.", "No event is lost or duplicated and 99% of local events register within 500 ms."),
            ("Acknowledge the final incident and inspect history.", "Acknowledgement does not remove any contributing event."),
        ],
        "expected": "Pass if all concurrent events are retained exactly once, one safe Alarm state is maintained, PR-01 timing passes, and the incident remains understandable. Fail on loss, duplicate side effects, crash, deadlock, or suppressed alarm.",
        "req": "FR-08, FR-15, PR-01, PR-05, IF-05",
    },
    {
        "id": "TC-LT-002",
        "suite": "Validation - Load Testing",
        "name": "High-volume User Access",
        "priority": "High",
        "type": "Non-functional — Performance, load, and security",
        "objective": "Validate consistent, authorized behavior when three or more devices access and interact with the application simultaneously.",
        "pre": "Five authenticated mobile/browser sessions across owner/member/guest roles; PR-05 maximum device inventory simulated; stable WAN; monitoring enabled; known initial Disarmed state.",
        "steps": [
            ("Have five sessions simultaneously request status, device health, and recent incident metadata for 10 min.", "Responses remain attributable, authorized, consistent, and free of another user's session data."),
            ("At the same instant, owner requests Arm-Away, member refreshes status, and guest attempts a prohibited configuration change.", "Authorized arm command is serialized once; status converges; guest change is denied and audited."),
            ("Trigger an alarm while all five sessions continue polling/viewing.", "Local PR-01–PR-03 behavior is not breached; active alarm/health state converges across sessions."),
            ("Review resource, error, and security logs.", "No crash, deadlock, privilege escalation, data leak, or unexplained error occurs."),
        ],
        "expected": "Pass if five simultaneous sessions operate within PR-05, authorized state changes are consistent/idempotent, prohibited access fails, and alarm timing remains compliant. Fail on stale unsafe state, leakage, duplicate command, or instability.",
        "req": "PR-01, PR-03, PR-05, IF-05, SEC-01, SEC-08",
    },
    {
        "id": "TC-LT-003",
        "suite": "Validation - Resilience",
        "name": "Internet Outage During Intrusion",
        "priority": "Critical",
        "type": "Non-functional — Reliability and resilience",
        "objective": "Validate the core local-first user need: protection and evidence continue during Internet loss and delayed alerts synchronize safely after recovery.",
        "pre": "Armed-Away; healthy sensor, CCU, siren, local panel, and associated camera; queued notifications empty; owner able to observe local and remote behavior; WAN disconnect controllable.",
        "steps": [
            ("Disconnect WAN and confirm the degraded indicator.", "Local UI identifies remote-connectivity loss without changing Armed-Away."),
            ("Trigger an enabled intrusion sensor.", "Local Alarm, siren, incident logging, and camera clip preservation occur without cloud service."),
            ("Attempt local authorized acknowledgement/disarm.", "Local control succeeds and retains the incident/evidence."),
            ("Restore WAN and observe synchronization.", "Queued alert and metadata transmit in chronological order and are visibly marked delayed; no duplicate local alarm occurs."),
            ("Review the incident with the owner.", "Owner can understand when the event occurred, when it was delivered, and which functions were degraded."),
        ],
        "expected": "Pass if all required local security functions continue offline, local control remains available, and recovery sends one ordered delayed alert with clear timestamps. Fail if WAN loss disarms/suppresses alarm, loses evidence, or causes duplicate/ambiguous recovery.",
        "req": "FR-13, FR-14, HS-01, RE-02, IF-06, VAL-02",
    },
]


def format_steps(items):
    return "\n".join(f"{index}. {action}\nExpected: {expected}" for index, (action, expected) in enumerate(items, 1))


def write_suite(sheet, cases, default_suite=None):
    headers = ["Test ID", "Suite", "Test Name", "Priority", "Type", "Objective", "Preconditions", "Steps", "Expected Result", "Requirements"]
    for col, value in enumerate(headers, 1):
        sheet.cell(1, col, value)

    for row, case in enumerate(cases, 2):
        values = [
            case["id"], case.get("suite", default_suite), case["name"], case["priority"], case["type"],
            case["objective"], case["pre"], format_steps(case["steps"]), case["expected"], case["req"],
        ]
        for col, value in enumerate(values, 1):
            sheet.cell(row, col, value)

    # Remove unused sample rows, if any.
    if sheet.max_row > len(cases) + 1:
        sheet.delete_rows(len(cases) + 2, sheet.max_row - len(cases) - 1)

    navy = "17365D"
    blue = "D9EAF7"
    white = "FFFFFF"
    thin = Side(style="thin", color="B7B7B7")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(bold=True, color=white)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in sheet.iter_rows(min_row=2, max_row=len(cases) + 1, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=blue)
    widths = [14, 27, 38, 12, 28, 55, 65, 95, 70, 32]
    for col, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.row_dimensions[1].height = 32
    for row in range(2, len(cases) + 2):
        sheet.row_dimensions[row].height = 170
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{len(cases) + 1}"
    sheet.sheet_view.showGridLines = False
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:1"


def main():
    copy2(SOURCE, OUTPUT)
    workbook = load_workbook(OUTPUT)
    write_suite(workbook["Verification Tests"], VERIFICATION, "Verification")
    write_suite(workbook["Validation Tests"], VALIDATION)
    workbook.save(OUTPUT)

    # Save/reopen validation.
    checked = load_workbook(OUTPUT, data_only=False)
    assert checked.sheetnames == ["Verification Tests", "Validation Tests"]
    assert checked["Verification Tests"].max_row == 8
    assert checked["Validation Tests"].max_row == 8
    for sheet in checked:
        assert sheet.max_column == 10
        for row in sheet.iter_rows(min_row=2, values_only=True):
            assert all(value not in (None, "") for value in row)
    checked.close()
    print(f"Completed {OUTPUT} with {len(VERIFICATION)} verification and {len(VALIDATION)} validation cases")


if __name__ == "__main__":
    main()
