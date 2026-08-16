"""Generate the Smart Home Security System weighted decision-matrix workbook."""

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


OUTPUT = Path(__file__).resolve().parents[1] / "decision-matrices.xlsx"

NAVY = "17365D"
BLUE = "D9EAF7"
LIGHT_BLUE = "EAF3F8"
GREEN = "E2F0D9"
GOLD = "FFF2CC"
RED = "F4CCCC"
WHITE = "FFFFFF"
GRAY = "666666"
THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


MATRICES = [
    {
        "sheet": "Sensor - Motion",
        "decision": "Select the motion-detection technology for protected indoor zones.",
        "criteria": [
            ("Detection accuracy", 20, "True human-motion detection across the protected area"),
            ("False-alarm resistance", 20, "Resistance to pets, HVAC, sunlight, curtains, and interference"),
            ("Battery life", 15, "Expected service interval for a battery-powered device"),
            ("Detection range / coverage", 10, "Useful distance and field of view"),
            ("Latency", 10, "Speed of reporting a valid event to the CCU"),
            ("Privacy", 10, "Minimizes collection of identifying or behavioral information"),
            ("Installation simplicity", 5, "Placement, calibration, and commissioning effort"),
            ("Unit + lifecycle cost", 5, "Purchase, batteries, maintenance, and replacement"),
            ("Interference resilience", 5, "Reliable operation around radio/electrical/environmental noise"),
        ],
        "options": [
            ("Passive infrared (PIR)", [4, 4, 5, 4, 5, 5, 5, 5, 4], "Best overall: mature, private, low-power, and fast; use pet-aware placement."),
            ("Microwave radar", [5, 2, 3, 5, 5, 4, 3, 3, 2], "Excellent range but greater through-wall/interference false-alarm risk."),
            ("Dual technology (PIR + microwave)", [5, 5, 3, 4, 4, 5, 3, 2, 5], "Highest confidence for critical zones, with higher cost and power demand."),
        ],
        "recommendation": "Use PIR as the standard indoor motion sensor; use dual-technology sensors in high-risk or false-alarm-prone zones.",
    },
    {
        "sheet": "Sensor - Entry",
        "decision": "Select the door/window opening sensor technology.",
        "criteria": [
            ("Open/close accuracy", 20, "Reliable state detection with low missed-event probability"),
            ("Tamper resistance", 15, "Resistance to removal, bypass, and magnetic defeat"),
            ("Battery life", 15, "Expected service interval"),
            ("Mounting flexibility", 10, "Works across frame materials, gaps, and door/window forms"),
            ("Latency", 10, "Speed of event delivery"),
            ("Reliability", 15, "Mechanical/electrical consistency over product life"),
            ("Installation simplicity", 5, "Time and skill required"),
            ("Unit + lifecycle cost", 5, "Purchase and maintenance burden"),
            ("Aesthetic impact", 5, "Visibility and effect on the opening"),
        ],
        "options": [
            ("Wireless magnetic reed contact", [5, 3, 5, 5, 5, 5, 5, 5, 4], "Best general retrofit option; pair with case/removal tamper detection."),
            ("Wired recessed magnetic contact", [5, 4, 5, 3, 5, 5, 2, 4, 5], "Best for new construction and critical openings; installation is invasive."),
            ("Wireless accelerometer/tilt sensor", [4, 4, 3, 5, 4, 4, 4, 3, 5], "Useful for unusual openings but needs calibration and more power."),
        ],
        "recommendation": "Use wireless magnetic contacts for retrofit installations and wired recessed contacts where cabling is available.",
    },
    {
        "sheet": "Sensor - Water",
        "decision": "Select the water-leak detection technology.",
        "criteria": [
            ("Detection coverage", 20, "Area/length over which escaping water is detected"),
            ("Detection speed", 15, "Time from water contact to reported event"),
            ("False-alarm resistance", 15, "Resistance to condensation, dirt, and incidental splashes"),
            ("Battery life", 10, "Expected service interval"),
            ("Placement flexibility", 10, "Suitability for appliances, pipes, basements, and irregular spaces"),
            ("Self-test / supervision", 10, "Ability to expose disconnection, battery, and sensor health"),
            ("Durability", 10, "Corrosion and environmental resistance"),
            ("Installation simplicity", 5, "Effort to place and commission"),
            ("Unit + lifecycle cost", 5, "Purchase and replacement cost"),
        ],
        "options": [
            ("Point probe / puck", [2, 5, 4, 5, 4, 4, 4, 5, 5], "Low-cost appliance protection; requires good placement at likely leak point."),
            ("Sensing cable / rope", [5, 5, 4, 4, 5, 5, 4, 3, 3], "Best overall coverage around tanks, appliances, and vulnerable pipe runs."),
            ("Flow-meter anomaly sensor", [5, 3, 3, 3, 3, 4, 5, 2, 2], "Whole-home coverage and hidden-leak potential, but needs plumbing integration."),
        ],
        "recommendation": "Use sensing cable in high-consequence zones, point probes under individual appliances, and treat flow monitoring as complementary whole-home protection.",
    },
    {
        "sheet": "Cameras",
        "decision": "Select the primary security-camera architecture.",
        "criteria": [
            ("Video reliability", 15, "Continuity and frame delivery under normal operation"),
            ("Image quality", 10, "Resolution, low-light performance, and evidence usefulness"),
            ("Outage resilience", 15, "Operation during WAN and/or local power disturbance"),
            ("Cybersecurity", 15, "Attack surface, updateability, authentication, and encryption"),
            ("Privacy", 10, "Local control and minimization of external video exposure"),
            ("Installation effort", 10, "Cabling, placement, and commissioning effort"),
            ("Bandwidth efficiency", 5, "Demand on Wi-Fi/WAN and predictability"),
            ("Scalability", 5, "Ease of adding cameras without contention or rework"),
            ("Power continuity", 5, "Ease of centralized backup power"),
            ("Lifecycle cost", 10, "Hardware, cabling, storage, subscriptions, and upkeep"),
        ],
        "options": [
            ("Wired PoE IP camera + local storage", [5, 5, 5, 5, 5, 2, 5, 4, 5, 3], "Recommended primary architecture: reliable, private, and centrally backed up."),
            ("Wi-Fi mains-powered IP camera", [3, 4, 4, 3, 4, 5, 3, 3, 3, 4], "Good retrofit option where Ethernet is impractical; depends on RF quality."),
            ("Battery Wi-Fi camera", [2, 3, 3, 3, 4, 5, 4, 3, 2, 4], "Fast installation but wake latency and battery servicing reduce assurance."),
            ("Cloud-managed camera", [3, 4, 1, 3, 1, 5, 1, 5, 2, 2], "Convenient remote access, with WAN, privacy, subscription, and lock-in tradeoffs."),
        ],
        "recommendation": "Use wired PoE cameras with encrypted local event storage; allow mains-powered Wi-Fi cameras as a retrofit exception.",
    },
    {
        "sheet": "Controllers",
        "decision": "Select the local connectivity platform for locks, lights, sirens, and relay controllers.",
        "criteria": [
            ("Command reliability", 20, "Probability that authorized commands execute and acknowledge"),
            ("Security", 20, "Secure enrollment, identity, encryption, and replay resistance"),
            ("Local/offline operation", 15, "Continues without Internet or vendor cloud"),
            ("Latency", 10, "Time to execute alarm and user commands"),
            ("Interoperability", 10, "Availability across lock, light, siren, and relay vendors"),
            ("Power efficiency", 5, "Suitability for battery-powered controllers"),
            ("Mesh/range", 5, "Coverage and resilience across the residence"),
            ("Installation simplicity", 5, "Enrollment and network setup effort"),
            ("Scalability", 5, "Device capacity and congestion behavior"),
            ("Lifecycle cost / lock-in", 5, "Hub, licensing, replacement, and vendor dependence"),
        ],
        "options": [
            ("Matter over Thread", [5, 5, 5, 5, 5, 5, 5, 4, 5, 4], "Recommended target platform: local IP interoperability with low-power mesh."),
            ("Zigbee 3.0", [5, 4, 5, 5, 4, 5, 5, 4, 5, 5], "Mature and efficient; profiles and vendor quirks need conformance testing."),
            ("Z-Wave", [5, 5, 5, 5, 4, 5, 5, 4, 4, 3], "Strong sub-GHz smart-home ecosystem but narrower sourcing/region considerations."),
            ("Wi-Fi / vendor cloud", [3, 3, 2, 4, 3, 2, 2, 5, 2, 2], "Simple per-device setup but higher power, contention, and cloud dependence."),
        ],
        "recommendation": "Adopt Matter over Thread as the strategic controller interface, with Zigbee 3.0 adapters for mature device coverage.",
    },
    {
        "sheet": "Server - Manager",
        "decision": "Select where the authoritative Smart Home System Manager executes.",
        "criteria": [
            ("Alarm availability", 20, "Ability to detect and alarm despite component or service outages"),
            ("Response latency", 10, "Time from local event to decision and actuator command"),
            ("Privacy / data control", 15, "Household control of recordings, events, and credentials"),
            ("Cybersecurity exposure", 10, "Internet attack surface and blast radius"),
            ("WAN independence", 10, "Operation through ISP/router/provider outages"),
            ("Remote access", 10, "Secure control and status outside the residence"),
            ("Scalability / analytics", 5, "Capacity growth and compute-intensive services"),
            ("Maintainability / updates", 5, "Deployment, diagnostics, backup, and recovery effort"),
            ("Lifecycle cost", 5, "Hardware, hosting, bandwidth, subscriptions, and support"),
            ("Disaster recovery", 5, "Recovery after local hardware loss or cloud-region failure"),
            ("Vendor independence", 5, "Portability and continued operation after provider exit"),
        ],
        "options": [
            ("Local edge server only", [5, 5, 5, 4, 5, 1, 2, 3, 4, 2, 5], "Strong local assurance and privacy, but weak remote access and off-site recovery."),
            ("Cloud server only", [1, 2, 2, 2, 1, 5, 5, 5, 3, 5, 2], "Operationally scalable but unsuitable as the sole authority for safety-critical local alarms."),
            ("Local-first hybrid", [5, 5, 4, 4, 5, 5, 4, 4, 3, 5, 4], "Recommended: local alarm authority with optional cloud relay, backup, and remote access."),
        ],
        "recommendation": "Select local-first hybrid: run the authoritative manager, state machine, database, queue, and evidence store on the CCU; use cloud services only for remote relay, notifications, and encrypted off-site metadata/backup.",
    },
]


def add_readme(workbook: Workbook) -> None:
    sheet = workbook.active
    sheet.title = "Guide"
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = "Smart Home Security System — Tradeoff Evaluation"
    sheet["A1"].font = Font(size=20, bold=True, color=WHITE)
    sheet["A1"].fill = PatternFill("solid", fgColor=NAVY)
    sheet.merge_cells("A1:F1")
    rows = [
        ("Purpose", "Compare architecture options using transparent, weighted criteria rather than cost alone."),
        ("Scoring", "1 = poor, 2 = below average, 3 = acceptable, 4 = good, 5 = excellent for this system context."),
        ("Weight", "Relative importance as a percentage. Weights on every matrix sum to 100%."),
        ("Weighted total", "SUMPRODUCT(weight, score) / SUM(weight), producing a total from 1.00 to 5.00."),
        ("Ranking", "1 is the highest weighted total. Scores are editable; formulas and rankings recalculate in Excel/Google Sheets."),
        ("Assumption", "Scores compare generic architectures, not vendor products. Validate vendor claims during procurement and site commissioning."),
        ("Decision rule", "The top score is the default selection. Any override should be documented with a constraint or risk treatment."),
        ("System context", "Residential local-first security system; local detection and alarming must survive loss of Internet service."),
    ]
    for row_idx, (label, value) in enumerate(rows, 3):
        sheet.cell(row_idx, 1, label).font = Font(bold=True, color=NAVY)
        sheet.cell(row_idx, 2, value)
        sheet.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=6)
        for col in range(1, 7):
            sheet.cell(row_idx, col).border = BORDER
            sheet.cell(row_idx, col).alignment = Alignment(vertical="top", wrap_text=True)
    sheet["A13"] = "Workbook sheets"
    sheet["A13"].font = Font(bold=True, color=WHITE)
    sheet["A13"].fill = PatternFill("solid", fgColor=NAVY)
    sheet.merge_cells("A13:F13")
    for idx, matrix in enumerate(MATRICES, 14):
        sheet.cell(idx, 1, matrix["sheet"])
        sheet.cell(idx, 2, matrix["decision"])
        sheet.merge_cells(start_row=idx, start_column=2, end_row=idx, end_column=6)
        sheet.cell(idx, 1).font = Font(bold=True)
        for col in range(1, 7):
            sheet.cell(idx, col).border = BORDER
            sheet.cell(idx, col).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 24
    for col in "BCDEF":
        sheet.column_dimensions[col].width = 20
    sheet.freeze_panes = "A3"


def add_matrix(workbook: Workbook, matrix: dict) -> None:
    sheet = workbook.create_sheet(matrix["sheet"])
    sheet.sheet_view.showGridLines = False
    option_count = len(matrix["options"])
    option_start = 4
    option_end = option_start + option_count - 1
    criteria_start = 3
    criteria_end = criteria_start + len(matrix["criteria"]) - 1
    total_col = criteria_end + 1
    rank_col = total_col + 1
    notes_col = rank_col + 1

    sheet.cell(1, 1, matrix["sheet"] + " Decision Matrix")
    sheet.cell(1, 1).font = Font(size=18, bold=True, color=WHITE)
    sheet.cell(1, 1).fill = PatternFill("solid", fgColor=NAVY)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=notes_col)
    sheet.cell(2, 1, "Decision: " + matrix["decision"])
    sheet.cell(2, 1).font = Font(bold=True, color=NAVY)
    sheet.cell(2, 1).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.cell(2, 1).comment = __import__("openpyxl").comments.Comment(matrix["decision"], "Architecture Team")

    sheet.cell(3, 1, "Option")
    sheet.cell(3, 2, "Weight →")
    for idx, (criterion, weight, definition) in enumerate(matrix["criteria"], criteria_start):
        sheet.cell(3, idx, criterion)
        sheet.cell(2, idx, weight / 100)
        sheet.cell(2, idx).number_format = "0%"
        sheet.cell(2, idx).comment = None
        sheet.cell(3, idx).alignment = Alignment(text_rotation=35, wrap_text=True, vertical="bottom", horizontal="center")
        sheet.cell(3, idx).comment = __import__("openpyxl").comments.Comment(definition, "Architecture Team")
    sheet.cell(3, total_col, "Weighted Total / 5")
    sheet.cell(3, rank_col, "Rank")
    sheet.cell(3, notes_col, "Tradeoff rationale")

    score_validation = DataValidation(type="whole", operator="between", formula1="1", formula2="5", allow_blank=False)
    score_validation.error = "Enter an integer score from 1 (poor) to 5 (excellent)."
    score_validation.errorTitle = "Invalid score"
    score_validation.prompt = "Score 1–5; see Guide for scale."
    score_validation.promptTitle = "Option score"
    score_validation.showInputMessage = True
    score_validation.showErrorMessage = True
    sheet.add_data_validation(score_validation)

    for row, (name, scores, rationale) in enumerate(matrix["options"], option_start):
        sheet.cell(row, 1, name)
        for idx, score in enumerate(scores, criteria_start):
            sheet.cell(row, idx, score)
        score_validation.add(f"{get_column_letter(criteria_start)}{row}:{get_column_letter(criteria_end)}{row}")
        weights = f"${get_column_letter(criteria_start)}$2:${get_column_letter(criteria_end)}$2"
        row_scores = f"{get_column_letter(criteria_start)}{row}:{get_column_letter(criteria_end)}{row}"
        sheet.cell(row, total_col, f"=SUMPRODUCT({weights},{row_scores})/SUM({weights})")
        totals = f"${get_column_letter(total_col)}${option_start}:${get_column_letter(total_col)}${option_end}"
        # RANK is supported by Excel, Google Sheets, and older spreadsheet engines.
        sheet.cell(row, rank_col, f"=RANK({get_column_letter(total_col)}{row},{totals},0)")
        sheet.cell(row, notes_col, rationale)
        sheet.cell(row, total_col).number_format = "0.00"
        sheet.cell(row, 1).font = Font(bold=True)

    rec_row = option_end + 2
    sheet.cell(rec_row, 1, "Recommendation")
    sheet.cell(rec_row, 1).font = Font(bold=True, color=NAVY)
    sheet.cell(rec_row, 2, matrix["recommendation"])
    sheet.merge_cells(start_row=rec_row, start_column=2, end_row=rec_row, end_column=notes_col)
    for col in range(1, notes_col + 1):
        sheet.cell(rec_row, col).fill = PatternFill("solid", fgColor=GREEN)
        sheet.cell(rec_row, col).border = BORDER
        sheet.cell(rec_row, col).alignment = Alignment(wrap_text=True, vertical="top")

    score_note_row = rec_row + 2
    sheet.cell(score_note_row, 1, "Score scale")
    sheet.cell(score_note_row, 2, "1 Poor | 2 Below average | 3 Acceptable | 4 Good | 5 Excellent")
    sheet.merge_cells(start_row=score_note_row, start_column=2, end_row=score_note_row, end_column=notes_col)

    for row in range(2, option_end + 1):
        for col in range(1, notes_col + 1):
            cell = sheet.cell(row, col)
            cell.border = BORDER
            if row >= option_start:
                cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center" if criteria_start <= col <= rank_col else "left")
    for col in range(1, notes_col + 1):
        sheet.cell(3, col).fill = PatternFill("solid", fgColor=NAVY)
        sheet.cell(3, col).font = Font(bold=True, color=WHITE)
    for col in range(criteria_start, criteria_end + 1):
        sheet.cell(2, col).fill = PatternFill("solid", fgColor=GOLD)
        sheet.cell(2, col).font = Font(bold=True)
        sheet.cell(2, col).alignment = Alignment(horizontal="center")
    sheet.cell(2, criteria_start - 1, f"=SUM({get_column_letter(criteria_start)}2:{get_column_letter(criteria_end)}2)")
    sheet.cell(2, criteria_start - 1).number_format = '0%;[Red]-0%'
    sheet.cell(2, criteria_start - 1).comment = __import__("openpyxl").comments.Comment("Weight check: must equal 100%.", "Architecture Team")

    score_range = f"{get_column_letter(criteria_start)}{option_start}:{get_column_letter(criteria_end)}{option_end}"
    sheet.conditional_formatting.add(score_range, ColorScaleRule(start_type="num", start_value=1, start_color=RED, mid_type="num", mid_value=3, mid_color=GOLD, end_type="num", end_value=5, end_color=GREEN))
    total_range = f"{get_column_letter(total_col)}{option_start}:{get_column_letter(total_col)}{option_end}"
    sheet.conditional_formatting.add(total_range, ColorScaleRule(start_type="min", start_color=RED, mid_type="percentile", mid_value=50, mid_color=GOLD, end_type="max", end_color=GREEN))

    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Weighted option totals"
    chart.x_axis.title = "Score (1–5)"
    chart.height = 6.5
    chart.width = 11
    chart.x_axis.scaling.min = 0
    chart.x_axis.scaling.max = 5
    data = Reference(sheet, min_col=total_col, min_row=3, max_row=option_end)
    cats = Reference(sheet, min_col=1, min_row=option_start, max_row=option_end)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    sheet.add_chart(chart, f"A{score_note_row + 3}")

    sheet.freeze_panes = f"{get_column_letter(criteria_start)}4"
    sheet.auto_filter.ref = f"A3:{get_column_letter(notes_col)}{option_end}"
    sheet.row_dimensions[3].height = 88
    sheet.row_dimensions[2].height = 48
    sheet.row_dimensions[rec_row].height = 52
    sheet.column_dimensions["A"].width = 34
    sheet.column_dimensions["B"].width = 12
    for col in range(criteria_start, criteria_end + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 13
    sheet.column_dimensions[get_column_letter(total_col)].width = 19
    sheet.column_dimensions[get_column_letter(rank_col)].width = 10
    sheet.column_dimensions[get_column_letter(notes_col)].width = 56
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.print_title_rows = "1:3"


def main() -> None:
    workbook = Workbook()
    add_readme(workbook)
    for matrix in MATRICES:
        add_matrix(workbook, matrix)
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.save(OUTPUT)

    # Reopen to catch malformed workbook structures before handoff.
    checked = load_workbook(OUTPUT, data_only=False)
    expected = ["Guide"] + [matrix["sheet"] for matrix in MATRICES]
    assert checked.sheetnames == expected
    for matrix in MATRICES:
        assert sum(weight for _, weight, _ in matrix["criteria"]) == 100
        assert all(len(scores) == len(matrix["criteria"]) for _, scores, _ in matrix["options"])
    checked.close()
    print(f"Created {OUTPUT} with {len(expected)} worksheets")


if __name__ == "__main__":
    main()
